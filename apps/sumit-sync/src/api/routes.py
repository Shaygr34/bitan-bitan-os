"""
FastAPI routes for IDOM-SUMIT Sync.

POST   /runs                         — create a new reconciliation run
POST   /runs/{id}/upload             — upload IDOM or SUMIT file
POST   /runs/{id}/execute            — run reconciliation engine
GET    /runs/{id}                    — get run detail (metrics, exceptions, files)
GET    /runs                         — list runs
DELETE /runs/{id}                    — delete run, its DB records, and stored files
GET    /runs/{id}/files/{fid}/download — download a file
GET    /runs/{id}/drill-down/{metric} — drill-down into metric data
PATCH  /runs/{id}/exceptions/{eid}   — update exception resolution
PATCH  /runs/{id}/exceptions/bulk    — bulk update exceptions
POST   /runs/{id}/complete           — mark run as completed (locks mutations)
"""

import time
import uuid as uuid_mod
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from ..db.connection import get_db
from ..db import models
from ..storage import file_store
from .schemas import (
    CreateRunRequest,
    PatchExceptionRequest,
    BulkPatchExceptionsRequest,
    RunOut,
    RunDetailOut,
    RunFileOut,
    RunMetricsOut,
    ExceptionOut,
    ExecuteResultOut,
    BulkPatchResultOut,
    DrillDownOut,
    MappingSummaryOut,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/runs", tags=["runs"])


def _to_uuid(value: str) -> uuid_mod.UUID:
    """Convert a string to uuid.UUID so SQLAlchemy Uuid columns work on SQLite."""
    try:
        return uuid_mod.UUID(value)
    except (ValueError, AttributeError):
        raise HTTPException(status_code=404, detail=f"מזהה לא תקין: {value}")


def _run_or_404(run_id: str, db: Session) -> models.Run:
    run = db.query(models.Run).filter(models.Run.id == _to_uuid(run_id)).first()
    if not run:
        raise HTTPException(status_code=404, detail=f"הרצה {run_id} לא נמצאה")
    return run


def _serialize_run(run: models.Run) -> RunOut:
    return RunOut(
        id=str(run.id),
        year=run.year,
        report_type=run.report_type,
        status=run.status,
        operator_notes=run.operator_notes,
        created_at=run.created_at,
        started_at=run.started_at,
        completed_at=run.completed_at,
    )


def _serialize_exception(e: models.Exception) -> ExceptionOut:
    return ExceptionOut(
        id=str(e.id),
        exception_type=e.exception_type,
        severity=e.severity,
        idom_ref=e.idom_ref,
        sumit_ref=e.sumit_ref,
        client_name=e.client_name,
        description=e.description,
        field_diffs=e.field_diffs,
        resolution=e.resolution,
        resolution_note=e.resolution_note,
        resolved_at=e.resolved_at,
        created_at=e.created_at,
    )


def _ensure_not_completed(run: models.Run):
    """Guard: block mutations on completed runs."""
    if run.status == "completed":
        raise HTTPException(409, "ההרצה הושלמה — לא ניתן לבצע שינויים")


def _serialize_run_detail(run: models.Run) -> RunDetailOut:
    files = [
        RunFileOut(
            id=str(f.id),
            file_role=f.file_role,
            original_name=f.original_name,
            size_bytes=f.size_bytes,
            uploaded_at=f.uploaded_at,
        )
        for f in run.files
    ]
    metrics = None
    if run.metrics:
        m = run.metrics
        metrics = RunMetricsOut(
            total_idom_records=m.total_idom_records,
            total_sumit_records=m.total_sumit_records,
            matched_count=m.matched_count,
            unmatched_count=m.unmatched_count,
            changed_count=m.changed_count,
            unchanged_count=m.unchanged_count,
            status_completed_count=m.status_completed_count,
            status_preserved_count=m.status_preserved_count,
            status_regression_flags=m.status_regression_flags,
            processing_seconds=m.processing_seconds,
        )
    exceptions = [_serialize_exception(e) for e in run.exceptions]
    return RunDetailOut(
        id=str(run.id),
        year=run.year,
        report_type=run.report_type,
        status=run.status,
        operator_notes=run.operator_notes,
        created_at=run.created_at,
        started_at=run.started_at,
        completed_at=run.completed_at,
        files=files,
        metrics=metrics,
        exceptions=exceptions,
    )


# ------------------------------------------------------------------ #
#  POST /runs  — create run
# ------------------------------------------------------------------ #

@router.post("", response_model=RunOut, status_code=201)
def create_run(body: CreateRunRequest, db: Session = Depends(get_db)):
    run = models.Run(
        year=body.year,
        report_type=body.report_type,
        status="uploading",
        operator_notes=body.operator_notes,
    )
    db.add(run)
    db.commit()
    db.refresh(run)
    logger.info("Created run %s (year=%d, type=%s)", run.id, run.year, run.report_type)
    return _serialize_run(run)


# ------------------------------------------------------------------ #
#  POST /runs/{id}/upload  — upload file
# ------------------------------------------------------------------ #

VALID_FILE_ROLES = {"idom_upload", "sumit_upload"}

@router.post("/{run_id}/upload", response_model=RunFileOut)
async def upload_file(
    run_id: str,
    file_role: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    run = _run_or_404(run_id, db)
    _ensure_not_completed(run)

    if run.status not in ("uploading", "review"):
        raise HTTPException(400, "לא ניתן להעלות קבצים בסטטוס הנוכחי")

    if file_role not in VALID_FILE_ROLES:
        raise HTTPException(400, f"תפקיד קובץ לא תקין. יש לבחור מתוך: {VALID_FILE_ROLES}")

    # Check for duplicate role
    existing = (
        db.query(models.RunFile)
        .filter(models.RunFile.run_id == run.id, models.RunFile.file_role == file_role)
        .first()
    )
    if existing:
        raise HTTPException(409, f"קובץ בתפקיד '{file_role}' כבר הועלה להרצה זו")

    content = await file.read()
    stored_path = file_store.store_upload(str(run.id), file.filename, content)

    run_file = models.RunFile(
        run_id=run.id,
        file_role=file_role,
        original_name=file.filename,
        stored_path=str(stored_path),
        size_bytes=len(content),
        mime_type=file.content_type,
    )
    db.add(run_file)
    db.commit()
    db.refresh(run_file)

    logger.info("Uploaded %s for run %s (%d bytes)", file_role, run_id, len(content))
    return RunFileOut(
        id=str(run_file.id),
        file_role=run_file.file_role,
        original_name=run_file.original_name,
        size_bytes=run_file.size_bytes,
        uploaded_at=run_file.uploaded_at,
    )


# ------------------------------------------------------------------ #
#  POST /runs/{id}/execute  — run reconciliation
# ------------------------------------------------------------------ #

@router.post("/{run_id}/execute", response_model=ExecuteResultOut)
def execute_run(run_id: str, db: Session = Depends(get_db)):
    run = _run_or_404(run_id, db)

    if run.status != "uploading":
        raise HTTPException(400, f"ניתן להריץ רק בסטטוס 'העלאה' (סטטוס נוכחי: {run.status})")

    # Verify both files are uploaded
    files_by_role = {f.file_role: f for f in run.files}
    if "idom_upload" not in files_by_role:
        raise HTTPException(400, "קובץ IDOM טרם הועלה")
    if "sumit_upload" not in files_by_role:
        raise HTTPException(400, "קובץ SUMIT טרם הועלה")

    # Transition to processing
    run.status = "processing"
    run.started_at = datetime.now(timezone.utc)
    db.commit()

    t0 = time.monotonic()

    try:
        result, output_paths, warnings = _run_reconciliation(
            idom_path=files_by_role["idom_upload"].stored_path,
            sumit_path=files_by_role["sumit_upload"].stored_path,
            report_type=run.report_type,
            tax_year=run.year,
            run_id=str(run.id),
        )
    except Exception as exc:
        run.status = "failed"
        run.completed_at = datetime.now(timezone.utc)
        db.commit()
        logger.exception("Reconciliation failed for run %s", run_id)
        raise HTTPException(500, f"הסנכרון נכשל: {exc}")

    elapsed = time.monotonic() - t0

    # Persist metrics
    metrics = models.RunMetrics(
        run_id=run.id,
        total_idom_records=result.total_idom_records,
        total_sumit_records=result.total_sumit_records,
        matched_count=result.matched_count,
        unmatched_count=result.unmatched_count,
        changed_count=result.changed_count,
        unchanged_count=result.unchanged_count,
        status_completed_count=result.status_completed_count,
        status_preserved_count=result.status_preserved_count,
        status_regression_flags=result.status_regression_flags,
        processing_seconds=round(elapsed, 3),
    )
    db.add(metrics)

    # Persist exceptions — unmatched records
    if not result.exceptions_df.empty:
        for _, exc_row in result.exceptions_df.iterrows():
            exc_record = models.Exception(
                run_id=run.id,
                exception_type=exc_row.get("exception_type", "no_sumit_match"),
                idom_ref=str(exc_row.get("מספר_תיק", "")),
                client_name=str(exc_row.get("שם", "")),
                description=str(exc_row.get("notes", "רשומת IDOM ללא התאמה בייצוא SUMIT. ייתכן לקוח חדש או סינון ייצוא.")),
            )
            db.add(exc_record)

    # Persist status regression exceptions — per record
    for rec in result.regression_records:
        exc_record = models.Exception(
            run_id=run.id,
            exception_type="status_regression",
            idom_ref=rec.get("idom_ref", ""),
            sumit_ref=rec.get("sumit_ref", ""),
            client_name=rec.get("client_name", ""),
            description=(
                "סטטוס 'הושלם' ב-SUMIT אך ללא תאריך הגשה ב-IDOM. "
                "הסטטוס נשמר — נדרשת בדיקה."
            ),
        )
        db.add(exc_record)

    # Persist IDOM duplicate exceptions (from parser conflicts)
    # These are passed through warnings; we handle them if present
    for w in warnings:
        if "duplicate" in w.lower() or "conflict" in w.lower():
            exc_record = models.Exception(
                run_id=run.id,
                exception_type="idom_duplicate",
                description=w,
            )
            db.add(exc_record)

    # Persist output file records
    output_file_names = []
    role_map = {
        "import": "import_output",
        "diff": "diff_report",
        "exceptions": "exceptions_report",
    }
    for key, path_str in output_paths.items():
        role = role_map.get(key, key)
        p = Path(path_str)
        run_file = models.RunFile(
            run_id=run.id,
            file_role=role,
            original_name=p.name,
            stored_path=path_str,
            size_bytes=p.stat().st_size if p.exists() else 0,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        db.add(run_file)
        output_file_names.append(p.name)

    # Finalize run status
    has_exceptions = result.unmatched_count > 0 or result.status_regression_flags > 0
    run.status = "review" if has_exceptions else "completed"
    run.completed_at = datetime.now(timezone.utc)
    db.commit()

    logger.info(
        "Run %s completed in %.2fs: %d matched, %d exceptions",
        run_id, elapsed, result.matched_count,
        result.unmatched_count + result.status_regression_flags,
    )

    return ExecuteResultOut(
        run_id=str(run.id),
        status=run.status,
        metrics=RunMetricsOut(
            total_idom_records=result.total_idom_records,
            total_sumit_records=result.total_sumit_records,
            matched_count=result.matched_count,
            unmatched_count=result.unmatched_count,
            changed_count=result.changed_count,
            unchanged_count=result.unchanged_count,
            status_completed_count=result.status_completed_count,
            status_preserved_count=result.status_preserved_count,
            status_regression_flags=result.status_regression_flags,
            processing_seconds=round(elapsed, 3),
        ),
        exception_count=result.unmatched_count + result.status_regression_flags,
        output_files=output_file_names,
        warnings=result.warnings,
    )


# ------------------------------------------------------------------ #
#  GET /runs/{id}  — run detail
# ------------------------------------------------------------------ #

@router.get("/{run_id}", response_model=RunDetailOut)
def get_run(run_id: str, db: Session = Depends(get_db)):
    run = _run_or_404(run_id, db)
    return _serialize_run_detail(run)


# ------------------------------------------------------------------ #
#  DELETE /runs/{id}  — delete run + files
# ------------------------------------------------------------------ #

@router.delete("/{run_id}", status_code=204)
def delete_run(run_id: str, db: Session = Depends(get_db)):
    run = _run_or_404(run_id, db)

    # Delete physical files from volume first
    freed = file_store.cleanup_run_files(str(run.id))
    logger.info("Cleaned up %d bytes for run %s", freed, run_id)

    # Cascade delete handles run_files, run_metrics, exceptions
    db.delete(run)
    db.commit()

    logger.info("Deleted run %s", run_id)
    return None


# ------------------------------------------------------------------ #
#  GET /runs  — list runs
# ------------------------------------------------------------------ #

@router.get("", response_model=List[RunOut])
def list_runs(
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    runs = (
        db.query(models.Run)
        .order_by(models.Run.created_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    return [_serialize_run(r) for r in runs]


# ------------------------------------------------------------------ #
#  GET /runs/{id}/files/{fid}/download  — download file
# ------------------------------------------------------------------ #

@router.get("/{run_id}/files/{file_id}/download")
def download_file(run_id: str, file_id: str, db: Session = Depends(get_db)):
    run = _run_or_404(run_id, db)
    file_record = (
        db.query(models.RunFile)
        .filter(models.RunFile.id == _to_uuid(file_id), models.RunFile.run_id == run.id)
        .first()
    )
    if not file_record:
        raise HTTPException(404, "הקובץ לא נמצא")
    path = Path(file_record.stored_path)
    if not path.exists():
        raise HTTPException(410, "הקובץ כבר לא קיים בשרת")
    return FileResponse(
        path=str(path),
        filename=file_record.original_name or path.name,
        media_type=file_record.mime_type or "application/octet-stream",
    )


# ------------------------------------------------------------------ #
#  GET /runs/{id}/drill-down/{metric}  — metric row-level data
# ------------------------------------------------------------------ #

VALID_METRICS = {
    "idom_records", "sumit_records", "matched",
    "unmatched", "changed", "status_completed", "regressions",
}

METRIC_FILE_MAP = {
    "matched": ("import_output", "ייבוא"),
    "changed": ("diff_report", "שינויים"),
    "status_completed": ("diff_report", "שינויי סטטוס"),
    "idom_records": ("idom_upload", None),
    "sumit_records": ("sumit_upload", None),
}

METRIC_EXCEPTION_MAP = {
    "unmatched": "no_sumit_match",
    "regressions": "status_regression",
}

EXCEPTION_TYPE_LABELS_HE = {
    "no_sumit_match": "ללא התאמה ב-SUMIT",
    "idom_duplicate": "כפילות IDOM",
    "status_regression": "נסיגת סטטוס",
}

RESOLUTION_LABELS_HE = {
    "pending": "ממתין",
    "acknowledged": "נבדק",
    "dismissed": "נדחה",
}


def _read_excel_rows(
    path: str, sheet_name: Optional[str], limit: int, offset: int
) -> DrillDownOut:
    """Read an Excel sheet and return paginated rows for drill-down."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return DrillDownOut(metric="", total_rows=0, columns=[], rows=[])

    headers = [
        str(h) if h is not None else f"col_{i}"
        for i, h in enumerate(all_rows[0])
    ]
    data_rows = all_rows[1:]
    total = len(data_rows)
    page = data_rows[offset : offset + limit]

    result_rows: List[Dict[str, Any]] = []
    for row in page:
        d: Dict[str, Any] = {}
        for i, val in enumerate(row):
            if i >= len(headers):
                break
            if val is None:
                d[headers[i]] = ""
            elif hasattr(val, "strftime"):
                d[headers[i]] = val.strftime("%d/%m/%Y")
            else:
                d[headers[i]] = val
        result_rows.append(d)

    return DrillDownOut(metric="", total_rows=total, columns=headers, rows=result_rows)


@router.get("/{run_id}/drill-down/{metric}", response_model=DrillDownOut)
def drill_down(
    run_id: str,
    metric: str,
    limit: int = Query(default=100, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
):
    """Return row-level data for a specific metric card."""
    run = _run_or_404(run_id, db)

    if metric not in VALID_METRICS:
        raise HTTPException(
            400, f"מדד לא תקין. אפשרויות: {', '.join(sorted(VALID_METRICS))}"
        )

    # --- DB-based: query exceptions table ---
    if metric in METRIC_EXCEPTION_MAP:
        exc_type = METRIC_EXCEPTION_MAP[metric]
        query = db.query(models.Exception).filter(
            models.Exception.run_id == run.id,
            models.Exception.exception_type == exc_type,
        )
        total = query.count()
        excs = query.offset(offset).limit(limit).all()

        columns = ["סוג", "מספר תיק", "שם", "תיאור", "סטטוס"]
        rows = [
            {
                "סוג": EXCEPTION_TYPE_LABELS_HE.get(e.exception_type, e.exception_type),
                "מספר תיק": e.idom_ref or "",
                "שם": e.client_name or "",
                "תיאור": e.description,
                "סטטוס": RESOLUTION_LABELS_HE.get(e.resolution, e.resolution),
            }
            for e in excs
        ]
        return DrillDownOut(
            metric=metric, total_rows=total, columns=columns, rows=rows
        )

    # --- File-based: read Excel sheet ---
    file_role, sheet_name = METRIC_FILE_MAP[metric]
    file_record = (
        db.query(models.RunFile)
        .filter(
            models.RunFile.run_id == run.id,
            models.RunFile.file_role == file_role,
        )
        .first()
    )
    if not file_record:
        raise HTTPException(404, f"קובץ {file_role} לא נמצא להרצה")

    fpath = Path(file_record.stored_path)
    if not fpath.exists():
        logger.error(
            "Drill-down file missing: role=%s, path=%s, run=%s",
            file_role, fpath, run_id,
        )
        raise HTTPException(
            410,
            f"הקובץ לא נמצא בשרת (role={file_role}, path={file_record.stored_path})",
        )

    result = _read_excel_rows(str(fpath), sheet_name, limit, offset)
    result.metric = metric
    return result


# ------------------------------------------------------------------ #
#  PATCH /runs/{id}/exceptions/bulk  — bulk update exceptions
#  (must be defined BEFORE the /{exception_id} route to avoid
#   FastAPI matching "bulk" as a path parameter)
# ------------------------------------------------------------------ #

@router.patch("/{run_id}/exceptions/bulk", response_model=BulkPatchResultOut)
def bulk_patch_exceptions(
    run_id: str,
    body: BulkPatchExceptionsRequest,
    db: Session = Depends(get_db),
):
    run = _run_or_404(run_id, db)
    _ensure_not_completed(run)
    now = datetime.now(timezone.utc)
    updated = (
        db.query(models.Exception)
        .filter(
            models.Exception.run_id == run.id,
            models.Exception.resolution == "pending",
        )
        .update({
            "resolution": body.resolution,
            "resolved_at": now,
        })
    )
    db.commit()
    logger.info("Bulk %s %d exceptions (run %s)", body.resolution, updated, run_id)
    return BulkPatchResultOut(updated_count=updated)


# ------------------------------------------------------------------ #
#  PATCH /runs/{id}/exceptions/{eid}  — update exception resolution
# ------------------------------------------------------------------ #

@router.patch("/{run_id}/exceptions/{exception_id}", response_model=ExceptionOut)
def patch_exception(
    run_id: str,
    exception_id: str,
    body: PatchExceptionRequest,
    db: Session = Depends(get_db),
):
    run = _run_or_404(run_id, db)
    _ensure_not_completed(run)
    exc = (
        db.query(models.Exception)
        .filter(models.Exception.id == _to_uuid(exception_id), models.Exception.run_id == run.id)
        .first()
    )
    if not exc:
        raise HTTPException(404, "חריג לא נמצא")
    exc.resolution = body.resolution
    if body.note is not None:
        exc.resolution_note = body.note
    exc.resolved_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(exc)
    logger.info("Exception %s → %s (run %s)", exception_id, body.resolution, run_id)
    return _serialize_exception(exc)


# ------------------------------------------------------------------ #
#  POST /runs/{id}/complete  — mark run as completed (lock)
# ------------------------------------------------------------------ #

@router.post("/{run_id}/complete", response_model=RunOut)
def complete_run(run_id: str, db: Session = Depends(get_db)):
    run = _run_or_404(run_id, db)
    if run.status == "completed":
        raise HTTPException(409, "ההרצה כבר הושלמה")
    if run.status != "review":
        raise HTTPException(
            400,
            f"ניתן להשלים רק הרצות בסטטוס 'בדיקה' (סטטוס נוכחי: {run.status})",
        )
    run.status = "completed"
    run.completed_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(run)
    logger.info("Run %s marked as completed", run_id)
    return _serialize_run(run)


# ------------------------------------------------------------------ #
#  Internal: run the reconciliation pipeline
# ------------------------------------------------------------------ #

def _run_reconciliation(
    idom_path: str,
    sumit_path: str,
    report_type: str,
    tax_year: int,
    run_id: str,
):
    """
    Orchestrates the full reconciliation pipeline using src/core/ modules.
    Returns (SyncResult, output_paths_dict, all_warnings).

    Zero business logic lives here — this is pure wiring.
    """
    from ..core.config import get_config
    from ..core.idom_parser import parse_idom_file
    from ..core.sumit_parser import parse_sumit_file
    from ..core.sync_engine import run_sync
    from ..core.output_writer import write_outputs

    config = get_config(report_type)

    # Parse inputs
    idom_df, idom_conflicts, idom_warnings = parse_idom_file(idom_path)
    sumit_df, sumit_lookup, sumit_warnings = parse_sumit_file(sumit_path, config, tax_year)

    # Run sync
    result = run_sync(idom_df, sumit_df, sumit_lookup, config, tax_year)

    # Write outputs to volume
    output_dir = str(file_store.outputs_dir(run_id))
    output_paths = write_outputs(result, config, output_dir, tax_year)

    # Collect all warnings
    all_warnings = idom_warnings + sumit_warnings + result.warnings
    if not idom_conflicts.empty:
        all_warnings.append(
            f"{len(idom_conflicts)} כפילויות IDOM זוהו"
        )

    return result, output_paths, all_warnings


# ------------------------------------------------------------------ #
#  POST /runs/{id}/execute-api  — run reconciliation with Summit API
# ------------------------------------------------------------------ #

@router.post("/{run_id}/execute-api")
def execute_run_api(run_id: str, db: Session = Depends(get_db)):
    """
    Run reconciliation using Summit API as SUMIT data source.
    Returns immediately — sync runs in background thread.
    Frontend polls GET /runs/{id} for status updates.
    """
    import threading

    run = _run_or_404(run_id, db)

    if run.status not in ("uploading", "review"):
        raise HTTPException(400, f"ניתן להריץ רק בסטטוס 'העלאה' או 'בדיקה' (סטטוס נוכחי: {run.status})")

    # Verify IDOM file is uploaded
    files_by_role = {f.file_role: f for f in run.files}
    if "idom_upload" not in files_by_role:
        raise HTTPException(400, "קובץ IDOM טרם הועלה")

    # Transition to processing
    run.status = "processing"
    run.started_at = datetime.now(timezone.utc)
    db.commit()

    # Resolve imports and capture values BEFORE spawning thread
    from ..db.connection import SessionLocal as _SessionLocal

    bg_run_id = str(run.id)
    bg_idom_path = files_by_role["idom_upload"].stored_path
    bg_report_type = run.report_type
    bg_tax_year = run.year

    def _background_sync():
        """Runs in a separate thread with its own DB session."""
        bg_db = None
        try:
            bg_db = _SessionLocal()
        except Exception as init_err:
            logger.error("BG thread DB init failed for %s: %s", bg_run_id, init_err)
            return
        t0 = time.monotonic()
        logger.info("BG sync started for run %s", bg_run_id)

        try:
            bg_run = bg_db.query(models.Run).filter(models.Run.id == _to_uuid(bg_run_id)).first()
            if not bg_run:
                logger.error("Background sync: run %s not found", bg_run_id)
                return

            result, output_paths, warnings = _run_reconciliation_api(
                idom_path=bg_idom_path,
                report_type=bg_report_type,
                tax_year=bg_tax_year,
                run_id=bg_run_id,
            )

            elapsed = time.monotonic() - t0

            # Persist metrics
            metrics = models.RunMetrics(
                run_id=bg_run.id,
                total_idom_records=result.total_idom_records,
                total_sumit_records=result.total_sumit_records,
                matched_count=result.matched_count,
                unmatched_count=result.unmatched_count,
                changed_count=result.changed_count,
                unchanged_count=result.unchanged_count,
                status_completed_count=result.status_completed_count,
                status_preserved_count=result.status_preserved_count,
                status_regression_flags=result.status_regression_flags,
                processing_seconds=round(elapsed, 3),
            )
            bg_db.add(metrics)

            # Persist exceptions
            if not result.exceptions_df.empty:
                for _, exc_row in result.exceptions_df.iterrows():
                    exc_record = models.Exception(
                        run_id=bg_run.id,
                        exception_type=exc_row.get("exception_type", "no_sumit_match"),
                        idom_ref=str(exc_row.get("מספר_תיק", "")),
                        client_name=str(exc_row.get("שם", "")),
                        description=str(exc_row.get("notes", "רשומת IDOM ללא התאמה בייצוא SUMIT.")),
                    )
                    bg_db.add(exc_record)

            for rec in result.regression_records:
                exc_record = models.Exception(
                    run_id=bg_run.id,
                    exception_type="status_regression",
                    idom_ref=rec.get("idom_ref", ""),
                    sumit_ref=rec.get("sumit_ref", ""),
                    client_name=rec.get("client_name", ""),
                    description="סטטוס 'הושלם' ב-SUMIT אך ללא תאריך הגשה ב-IDOM.",
                )
                bg_db.add(exc_record)

            for w in warnings:
                if "duplicate" in w.lower() or "conflict" in w.lower():
                    exc_record = models.Exception(
                        run_id=bg_run.id,
                        exception_type="idom_duplicate",
                        description=w,
                    )
                    bg_db.add(exc_record)

            # Persist output files
            role_map = {"import": "import_output", "diff": "diff_report", "exceptions": "exceptions_report"}
            for key, path_str in output_paths.items():
                role = role_map.get(key, key)
                p = Path(path_str)
                run_file = models.RunFile(
                    run_id=bg_run.id,
                    file_role=role,
                    original_name=p.name,
                    stored_path=path_str,
                    size_bytes=p.stat().st_size if p.exists() else 0,
                    mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                bg_db.add(run_file)

            has_exceptions = result.unmatched_count > 0 or result.status_regression_flags > 0
            bg_run.status = "review" if has_exceptions else "completed"
            bg_run.completed_at = datetime.now(timezone.utc)
            bg_db.commit()

            logger.info(
                "Background sync %s completed in %.1fs: %d matched, %d exceptions",
                bg_run_id, elapsed, result.matched_count,
                result.unmatched_count + result.status_regression_flags,
            )

        except Exception as exc:
            logger.exception("Background sync %s failed: %s", bg_run_id, exc)
            try:
                bg_run = bg_db.query(models.Run).filter(models.Run.id == _to_uuid(bg_run_id)).first()
                if bg_run:
                    bg_run.status = "failed"
                    bg_run.operator_notes = "שגיאה: {}".format(str(exc)[:500])
                    bg_run.completed_at = datetime.now(timezone.utc)
                    bg_db.commit()
            except Exception:
                logger.exception("Failed to update run status after error")
        finally:
            bg_db.close()

    # Launch background thread and return immediately
    thread = threading.Thread(target=_background_sync, daemon=True, name=f"sync-{bg_run_id[:8]}")
    thread.start()
    logger.info("BG sync thread launched for run %s", bg_run_id)
    import sys as _sys
    _sys.stdout.flush()
    _sys.stderr.flush()

    return {"run_id": bg_run_id, "status": "processing", "message": "הסנכרון הופעל ברקע"}


# ------------------------------------------------------------------ #
#  GET /mapping  — view client mapping summary
# ------------------------------------------------------------------ #

@router.get("/mapping/summary", response_model=MappingSummaryOut, tags=["mapping"])
def get_mapping_summary():
    """Get summary stats for the client ↔ company number mapping store."""
    from ..core.mapping_store import MappingStore
    store = MappingStore()
    return store.to_summary()


# ------------------------------------------------------------------ #
#  POST /mapping/refresh  — rebuild mapping from API
# ------------------------------------------------------------------ #

@router.post("/mapping/refresh", tags=["mapping"])
def refresh_mapping():
    """
    Rebuild the client mapping by fetching all clients from Summit API.
    This is a slow operation (takes several minutes due to rate limits).
    """
    from ..core.sumit_api_client import SummitAPIClient
    from ..core.mapping_store import MappingStore

    api = SummitAPIClient()
    store = MappingStore()

    # Fetch all client IDs
    client_ids = api.list_entities("557688522")  # לקוחות folder

    resolved = 0
    for cid in client_ids:
        if not store.has_client(cid):
            cn = api.get_client_company_number(cid)
            if cn:
                store.add(cid, cn)
                resolved += 1

    store.save()

    return {
        "total_clients": len(client_ids),
        "newly_resolved": resolved,
        "total_mappings": store.size,
        "api_calls": api.call_count,
    }


# ------------------------------------------------------------------ #
#  Internal: run reconciliation with API source
# ------------------------------------------------------------------ #

def _run_reconciliation_api(
    idom_path: str,
    report_type: str,
    tax_year: int,
    run_id: str,
):
    """
    Orchestrates reconciliation using Summit API as data source.
    Only requires IDOM file — SUMIT data comes from API.
    Supports both multi-sheet workbooks and single-sheet files.
    """
    from ..core.config import get_config
    from ..core.idom_parser import parse_idom_file
    from ..core.idom_workbook import parse_idom_workbook
    from ..core.sumit_api_source import fetch_sumit_data
    from ..core.sync_engine import run_sync
    from ..core.output_writer import write_outputs

    config = get_config(report_type)

    # Try multi-sheet workbook first, fall back to single-sheet parser
    try:
        wb_result = parse_idom_workbook(idom_path)
        # Find the sheet matching the requested report type
        matching = [s for s in wb_result.sheets if s.report_type == report_type and s.error is None and len(s.records) > 0]
        if matching:
            sheet = matching[0]
            idom_df = sheet.records
            idom_conflicts = sheet.conflicts
            idom_warnings = sheet.warnings
            logger.info("Workbook: using sheet '%s' (%d records) for %s", sheet.sheet_name, len(idom_df), report_type)
        elif wb_result.total_records > 0:
            # Workbook has data but not for this report type — use first non-empty sheet
            first_good = next((s for s in wb_result.sheets if s.error is None and len(s.records) > 0), None)
            if first_good:
                idom_df = first_good.records
                idom_conflicts = first_good.conflicts
                idom_warnings = first_good.warnings + [
                    "Sheet '{}' used (no exact match for report type '{}')".format(first_good.sheet_name, report_type)
                ]
                logger.info("Workbook: no '%s' sheet, using '%s' (%d records)", report_type, first_good.sheet_name, len(idom_df))
            else:
                raise ValueError("Workbook parsed but all sheets empty or errored")
        else:
            raise ValueError("Workbook parsed but no records found")
    except Exception as wb_err:
        # Fall back to original single-sheet parser
        logger.info("Workbook parse failed (%s), falling back to single-sheet parser", wb_err)
        idom_df, idom_conflicts, idom_warnings = parse_idom_file(idom_path)

    # Fetch SUMIT data from API (replaces parse_sumit_file)
    sumit_df, sumit_lookup, sumit_warnings = fetch_sumit_data(config, tax_year)

    # Run sync
    result = run_sync(idom_df, sumit_df, sumit_lookup, config, tax_year)

    # Write outputs
    output_dir = str(file_store.outputs_dir(run_id))
    output_paths = write_outputs(result, config, output_dir, tax_year)

    # Collect warnings
    all_warnings = idom_warnings + sumit_warnings + result.warnings
    if not idom_conflicts.empty:
        all_warnings.append(f"{len(idom_conflicts)} כפילויות IDOM זוהו")

    return result, output_paths, all_warnings


# ------------------------------------------------------------------ #
#  Write-back endpoints
# ------------------------------------------------------------------ #

def _build_write_plan_for_run(run: models.Run):
    """Helper: build write plan from a run's data."""
    files_by_role = {f.file_role: f for f in run.files}
    if "idom_upload" not in files_by_role:
        raise HTTPException(400, "קובץ IDOM לא נמצא")

    from ..core.config import get_config
    from ..core.idom_parser import parse_idom_file
    from ..core.sumit_api_source import fetch_sumit_data
    from ..core.sync_engine import SyncEngine
    from ..core.mapping_store import MappingStore
    from ..core.taxonomy import load_full_taxonomies, is_loaded
    from ..core.sumit_api_client import SummitAPIClient

    config = get_config(run.report_type)
    idom_df, _, _ = parse_idom_file(files_by_role["idom_upload"].stored_path)
    sumit_df, sumit_lookup, _ = fetch_sumit_data(config, run.year)
    mapping = MappingStore()

    # Ensure full taxonomy tables are loaded
    if not is_loaded():
        try:
            api = SummitAPIClient()
            load_full_taxonomies(api)
        except Exception as e:
            logger.warning("Could not load full taxonomies: %s", e)

    engine = SyncEngine(config)
    return engine.build_write_plan(idom_df, sumit_df, sumit_lookup, run.year, mapping)


def _save_write_logs(run_id, audit_log, db: Session):
    """Persist audit log entries to DB."""
    for entry in audit_log:
        log = models.WriteLog(
            run_id=run_id,
            op_type=entry.get("op_type", ""),
            entity_id=entry.get("entity_id"),
            folder_id=entry.get("folder_id", ""),
            match_key=entry.get("match_key", ""),
            client_name=entry.get("client_name", ""),
            properties_written=entry.get("properties_written"),
            old_values=entry.get("old_values"),
            status=entry.get("status", ""),
            error_message=entry.get("error", ""),
        )
        db.add(log)
    db.commit()


@router.get("/{run_id}/write-plan", tags=["write-back"])
def get_write_plan(run_id: str, db: Session = Depends(get_db)):
    """Generate write plan for a completed sync run."""
    run = _run_or_404(run_id, db)
    if run.status not in ("review", "completed"):
        raise HTTPException(400, "תוכנית כתיבה דורשת הרצת סנכרון שהושלמה")

    plan = _build_write_plan_for_run(run)

    return {
        "summary": plan.summary(),
        "operations": [op.to_dict() for op in plan.operations],
    }


@router.post("/{run_id}/write-back/dry-run", tags=["write-back"])
def write_back_dry_run(run_id: str, db: Session = Depends(get_db)):
    """Execute write plan in dry-run mode (validates without writing)."""
    run = _run_or_404(run_id, db)
    if run.status not in ("review", "completed"):
        raise HTTPException(400, "כתיבה חוזרת דורשת הרצת סנכרון שהושלמה")

    plan = _build_write_plan_for_run(run)

    from ..core.write_executor import WriteExecutor
    executor = WriteExecutor(dry_run=True)
    result = executor.execute(plan)

    _save_write_logs(run.id, result.audit_log, db)

    return result.to_dict()


@router.post("/{run_id}/write-back", tags=["write-back"])
def write_back_live(run_id: str, db: Session = Depends(get_db)):
    """Execute write plan LIVE — writes to Summit CRM. Use with caution."""
    run = _run_or_404(run_id, db)
    if run.status not in ("review", "completed"):
        raise HTTPException(400, "כתיבה חוזרת דורשת הרצת סנכרון שהושלמה")

    plan = _build_write_plan_for_run(run)

    from ..core.write_executor import WriteExecutor
    executor = WriteExecutor(dry_run=False)
    result = executor.execute(plan)

    _save_write_logs(run.id, result.audit_log, db)

    return result.to_dict()
