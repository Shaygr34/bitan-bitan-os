"""
IDOM→SUMIT Sync Engine - Output Writer
Generates Excel output files: import, diff report, exceptions.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional
from datetime import datetime
from pathlib import Path
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .config import ReportConfig, STATUS_COMPLETED
from .sync_engine import SyncResult

logger = logging.getLogger(__name__)


class OutputWriter:
    """
    Writer for sync output files.
    """
    
    # Styling constants
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    CHANGE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ERROR_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, config: ReportConfig, output_dir: str):
        self.config = config
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    def write_all(self, result: SyncResult, tax_year: int) -> Dict[str, str]:
        """
        Write all output files.
        
        Returns:
            Dictionary of output file paths
        """
        report_type = self.config.report_type.value
        
        paths = {}
        
        # 1. SUMIT Import file
        import_path = self._write_import_file(result.import_df, tax_year)
        paths['import'] = str(import_path)
        
        # 2. Diff report
        diff_path = self._write_diff_report(result, tax_year)
        paths['diff'] = str(diff_path)
        
        # 3. Exceptions report
        exceptions_path = self._write_exceptions_report(result, tax_year)
        paths['exceptions'] = str(exceptions_path)
        
        logger.info(f"Output files written to {self.output_dir}")
        return paths
    
    def _write_import_file(self, df: pd.DataFrame, tax_year: int) -> Path:
        """Write SUMIT import file (values only)."""
        report_type = self.config.report_type.value
        filename = f"sumit_import_{report_type}_{tax_year}_{self.timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        if df.empty:
            # Create empty file with headers only
            wb = Workbook()
            ws = wb.active
            ws.title = "ייבוא"

            for col_idx, header in enumerate(self.config.import_schema.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL

            wb.save(filepath)
            logger.info(f"Written empty import file: {filepath}")
            return filepath

        # Clean data for export
        export_df = self._prepare_for_export(df)

        # Write to Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "ייבוא"
        
        # Write headers
        for col_idx, header in enumerate(export_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        
        # Write data (values only - no formulas)
        for row_idx, row in enumerate(export_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
        
        # Auto-width columns
        self._auto_width(ws)
        
        wb.save(filepath)
        logger.info(f"Written import file: {filepath} ({len(df)} records)")
        return filepath
    
    def _write_diff_report(self, result: SyncResult, tax_year: int) -> Path:
        """Write diff report with multiple sheets."""
        report_type = self.config.report_type.value
        filename = f"diff_report_{report_type}_{tax_year}_{self.timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "סיכום"
        self._write_summary_sheet(ws_summary, result, tax_year)

        # Sheet 2: All Changes
        ws_changes = wb.create_sheet("שינויים")
        self._write_changes_sheet(ws_changes, result.diff_df)

        # Sheet 3: Status Changes
        ws_status = wb.create_sheet("שינויי סטטוס")
        status_changes = result.diff_df[result.diff_df['field'] == 'סטטוס'] if not result.diff_df.empty else pd.DataFrame()
        self._write_changes_sheet(ws_status, status_changes)

        # Sheet 4: Extension Updates
        ws_ext = wb.create_sheet("עדכוני ארכה")
        ext_changes = result.diff_df[result.diff_df['field'] == 'אורכה משרד'] if not result.diff_df.empty else pd.DataFrame()
        self._write_changes_sheet(ws_ext, ext_changes)

        # Sheet 5: Warnings
        ws_warnings = wb.create_sheet("אזהרות")
        self._write_warnings_sheet(ws_warnings, result.warnings)
        
        wb.save(filepath)
        logger.info(f"Written diff report: {filepath}")
        return filepath
    
    def _write_exceptions_report(self, result: SyncResult, tax_year: int) -> Path:
        """Write exceptions report."""
        report_type = self.config.report_type.value
        filename = f"exceptions_{report_type}_{tax_year}_{self.timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        wb = Workbook()
        
        # Sheet 1: Unmatched Records
        ws_unmatched = wb.active
        ws_unmatched.title = "ללא התאמה"
        self._write_dataframe_sheet(ws_unmatched, result.exceptions_df, "רשומות IDOM ללא התאמה")
        
        # Sheet 2: Status Regression Flags
        ws_regression = wb.create_sheet("סקירת סטטוס")
        if result.status_regression_flags > 0:
            if result.regression_records:
                regression_df = pd.DataFrame(result.regression_records)
                regression_df.columns = [
                    col.replace("idom_ref", "מספר תיק")
                       .replace("sumit_ref", "מזהה SUMIT")
                       .replace("client_name", "שם")
                       .replace("status", "סטטוס")
                    for col in regression_df.columns
                ]
                self._write_dataframe_sheet(ws_regression, regression_df, "נסיגות סטטוס")
            else:
                regression_note = pd.DataFrame([{
                    'הערה': f'{result.status_regression_flags} רשומות עם סטטוס "הושלם" ב-SUMIT אך ללא תאריך הגשה ב-IDOM.',
                    'פעולה': 'הסטטוס נשמר כהושלם. נדרשת בדיקה.'
                }])
                self._write_dataframe_sheet(ws_regression, regression_note, "נסיגות סטטוס")
        else:
            ws_regression.cell(row=1, column=1, value="אין נסיגות סטטוס")
        
        # Sheet 3: Summary
        ws_summary = wb.create_sheet("סיכום")
        summary_data = pd.DataFrame([
            {'מדד': 'רשומות IDOM', 'ערך': result.total_idom_records},
            {'מדד': 'רשומות SUMIT', 'ערך': result.total_sumit_records},
            {'מדד': 'התאמות', 'ערך': result.matched_count},
            {'מדד': 'ללא התאמה (חריגים)', 'ערך': result.unmatched_count},
            {'מדד': 'שינויים', 'ערך': result.changed_count},
            {'מדד': 'נסיגות סטטוס', 'ערך': result.status_regression_flags},
        ])
        self._write_dataframe_sheet(ws_summary, summary_data, "סיכום חריגים")
        
        wb.save(filepath)
        logger.info(f"Written exceptions report: {filepath}")
        return filepath
    
    def _write_summary_sheet(self, ws, result: SyncResult, tax_year: int) -> None:
        """Write summary sheet."""
        # Title
        ws.cell(row=1, column=1, value="דו״ח סנכרון IDOM → SUMIT")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        ws.cell(row=2, column=1, value=f"סוג דו״ח: {self.config.display_name}")
        ws.cell(row=3, column=1, value=f"שנת מס: {tax_year}")
        ws.cell(row=4, column=1, value=f"נוצר: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        # Stats table
        stats = [
            ("", ""),
            ("סטטיסטיקת עיבוד", ""),
            ("רשומות IDOM", result.total_idom_records),
            ("רשומות SUMIT", result.total_sumit_records),
            ("", ""),
            ("תוצאות התאמה", ""),
            ("התאמות", result.matched_count),
            ("ללא התאמה (חריגים)", result.unmatched_count),
            ("אחוז התאמה", f"{result.matched_count / max(result.total_idom_records, 1) * 100:.1f}%"),
            ("", ""),
            ("סטטיסטיקת עדכונים", ""),
            ("רשומות שהשתנו", result.changed_count),
            ("רשומות ללא שינוי", result.unchanged_count),
            ("סטטוס → הושלם", result.status_completed_count),
            ("סטטוס נשמר", result.status_preserved_count),
            ("נסיגות סטטוס", result.status_regression_flags),
        ]

        for row_idx, (label, value) in enumerate(stats, 6):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)

            if label in ["סטטיסטיקת עיבוד", "תוצאות התאמה", "סטטיסטיקת עדכונים"]:
                ws.cell(row=row_idx, column=1).font = Font(bold=True)

        # Warnings section
        if result.warnings:
            warning_row = len(stats) + 8
            ws.cell(row=warning_row, column=1, value="אזהרות")
            ws.cell(row=warning_row, column=1).font = Font(bold=True)
            
            for idx, warning in enumerate(result.warnings, 1):
                ws.cell(row=warning_row + idx, column=1, value=warning)
                ws.cell(row=warning_row + idx, column=1).fill = self.CHANGE_FILL
        
        self._auto_width(ws)
    
    def _write_changes_sheet(self, ws, df: pd.DataFrame) -> None:
        """Write changes DataFrame to sheet."""
        if df.empty:
            ws.cell(row=1, column=1, value="לא נרשמו שינויים")
            return

        self._write_dataframe_sheet(ws, df, "שינויים")
    
    def _write_warnings_sheet(self, ws, warnings: List[str]) -> None:
        """Write warnings to sheet."""
        ws.cell(row=1, column=1, value="אזהרות")
        ws.cell(row=1, column=1).font = Font(bold=True)

        if not warnings:
            ws.cell(row=2, column=1, value="אין אזהרות")
            return
        
        for idx, warning in enumerate(warnings, 2):
            ws.cell(row=idx, column=1, value=warning)
    
    def _write_dataframe_sheet(self, ws, df: pd.DataFrame, title: str = "") -> None:
        """Write DataFrame to worksheet."""
        if df.empty:
            ws.cell(row=1, column=1, value=f"אין נתונים ({title})" if title else "אין נתונים")
            return
        
        # Headers
        for col_idx, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=self._clean_value(value))
                cell.border = self.THIN_BORDER
        
        self._auto_width(ws)
    
    def _prepare_for_export(self, df: pd.DataFrame) -> pd.DataFrame:
        """Prepare DataFrame for Excel export (values only)."""
        export_df = df.copy()
        
        for col in export_df.columns:
            export_df[col] = export_df[col].apply(self._clean_value)
        
        return export_df
    
    @staticmethod
    def _clean_value(value) -> Any:
        """Clean value for Excel export."""
        if pd.isna(value):
            return ''
        
        if isinstance(value, (datetime, pd.Timestamp)):
            # SUMIT requires dd/MM/yyyy format
            return value.strftime('%d/%m/%Y')
        
        # Handle float .0 suffix
        if isinstance(value, float) and value == int(value):
            return int(value)
        
        return value
    
    @staticmethod
    def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, cell_length)
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width


def write_outputs(
    result: SyncResult,
    config: ReportConfig,
    output_dir: str,
    tax_year: int
) -> Dict[str, str]:
    """
    Convenience function to write all outputs.
    """
    writer = OutputWriter(config, output_dir)
    return writer.write_all(result, tax_year)
